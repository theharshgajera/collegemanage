from django.shortcuts import render, redirect
from django.http import HttpResponse
from app.models import Student_Notification, Student, Student_Feedback, Student_leave, Subject, Attendance, Attendance_Report, StudentResult, Session_Year, Note, StudentDoubt
from django.contrib import messages
from django.contrib.auth.decorators import login_required
import openpyxl
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib.styles import getSampleStyleSheet
import io
import google.generativeai as genai
from django.conf import settings

@login_required(login_url='/')
# Remove this function as it's no longer needed
# @login_required(login_url='/')
# def calculate_attendance_score(present_count, total_count, max_score=10):
#     if total_count == 0:
#         return 0
#     attendance_percentage = (present_count / total_count) * 100
#     attendance_score = (attendance_percentage / 100) * max_score
#     return round(attendance_score, 2)

@login_required(login_url='/')
def HOME(request):
    try:
        # Check if user is a student
        if request.user.user_type != '3':
            messages.error(request, "Access Denied: You are not a student")
            return redirect('login')
            
        student = Student.objects.get(admin=request.user)
        
        # Get total subjects
        total_subjects = Subject.objects.filter(course=student.course_id).count()
        
        # Calculate attendance percentage
        attendance_present = Attendance_Report.objects.filter(
            student_id=student,
            status=1
        ).count()
        total_attendance = Attendance_Report.objects.filter(student_id=student).count()
        attendance_percentage = (attendance_present / total_attendance * 100) if total_attendance > 0 else 0
        
        context = {
            'student': student,
            'total_subjects': total_subjects,
            'attendance_percentage': round(attendance_percentage, 2),
            'notes_count': Note.objects.filter(user=request.user).count(),
            'doubts_count': StudentDoubt.objects.filter(student=student).count(),
            'leave_count': Student_leave.objects.filter(student_id=student).count(),
            'pending_leaves': Student_leave.objects.filter(student_id=student, status=0).count(),
            'results': StudentResult.objects.filter(student_id=student),
        }
        return render(request, 'Student/home.html', context)
        
    except Student.DoesNotExist:
        messages.error(request, "Student record not found")
        return redirect('login')
    except Exception as e:
        messages.error(request, f"An error occurred: {str(e)}")
        return redirect('login')

@login_required(login_url='/')
def STUDENT_NOTIFICATION(request):
    student = Student.objects.filter(admin=request.user.id)
    for i in student:
        student_id = i.id
        notification = Student_Notification.objects.filter(student_id=student_id)
        context = {'notification': notification}
        return render(request, 'Student/notification.html', context)

@login_required(login_url='/')
def STUDENT_NOTIFICATION_MARK_AS_DONE(request, status):
    notification = Student_Notification.objects.get(id=status)
    notification.status = 1
    notification.save()
    return redirect('student_notification')

@login_required(login_url='/')
def STUDENT_FEEDBACK(request):
    student_id = Student.objects.get(admin=request.user.id)
    feedback_history = Student_Feedback.objects.filter(student_id=student_id).order_by('-id')
    context = {"feedback_history": feedback_history}
    return render(request, 'Student/feedback.html', context)

@login_required(login_url='/')
def STUDENT_FEEDBACK_SAVE(request):
    if request.method == "POST":
        feedback = request.POST.get('feedback')
        student = Student.objects.get(admin=request.user.id)
        feedbacks = Student_Feedback(student_id=student, feedback=feedback, feedback_reply="")
        feedbacks.save()
        messages.success(request, 'Feedback Successfully Sent!')
        return redirect('student_feedback')
    return redirect('student_feedback')

@login_required(login_url='/')
def STUDENT_LEAVE(request):
    student = Student.objects.get(admin=request.user.id)
    student_leave_history = Student_leave.objects.filter(student_id=student)
    context = {'student_leave_history': student_leave_history}
    return render(request, 'Student/apply_leave.html', context)

@login_required(login_url='/')
def STUDENT_LEAVE_SAVE(request):
    if request.method == "POST":
        leave_date = request.POST.get('leave_date')
        leave_message = request.POST.get('leave_message')

        if not leave_date:
            messages.error(request, 'Please provide a valid leave date.')
            return redirect('student_leave')
        if not leave_message or leave_message.strip() == '':
            messages.error(request, 'Please provide a reason for the leave.')
            return redirect('student_leave')

        student_id = Student.objects.get(admin=request.user.id)
        student_leave = Student_leave(student_id=student_id, date=leave_date, message=leave_message)
        student_leave.save()
        messages.success(request, 'Leave Application Successfully Sent!')
        return redirect('student_leave')
    return redirect('student_leave')

@login_required(login_url='/')
def STUDENT_VIEW_ATTENDANCE(request):
    student = Student.objects.get(admin=request.user.id)
    subjects = Subject.objects.filter(course=student.course_id)
    session_year = Session_Year.objects.all()
    action = request.GET.get('action')
    get_subject = None
    get_session_year = None
    attendance_report = None

    if action == 'show_attendance' and request.method == "POST":
        subject_id = request.POST.get('subject_id')
        session_year_id = request.POST.get('session_year_id')
        try:
            if subject_id and subject_id.isdigit() and session_year_id and session_year_id.isdigit():
                get_subject = Subject.objects.get(id=int(subject_id))
                get_session_year = Session_Year.objects.get(id=int(session_year_id))
                if get_subject.course != student.course_id or get_session_year.id != student.session_year_id.id:
                    messages.error(request, "You are not authorized to view attendance for this subject or session year.")
                    return redirect('student_view_attendance')
                attendance_report = Attendance_Report.objects.filter(
                    student_id=student,
                    attendance_id__subject_id=get_subject,
                    attendance_id__session_year_id=get_session_year
                ).order_by('attendance_id__attendance_data')
            else:
                messages.error(request, "Please select a valid subject and session year.")
                return redirect('student_view_attendance')
        except Subject.DoesNotExist:
            messages.error(request, "Selected subject does not exist.")
            return redirect('student_view_attendance')
        except Session_Year.DoesNotExist:
            messages.error(request, "Selected session year does not exist.")
            return redirect('student_view_attendance')

    if 'download' in request.GET and request.GET['download'] == 'excel':
        subject_id = request.GET.get('subject_id')
        session_year_id = request.GET.get('session_year_id')
        print(f"Download Excel triggered - subject_id: {subject_id}, session_year_id: {session_year_id}")

        try:
            get_subject = Subject.objects.get(id=int(subject_id))
            get_session_year = Session_Year.objects.get(id=int(session_year_id))
            if get_subject.course != student.course_id or get_session_year.id != student.session_year_id.id:
                messages.error(request, "You are not authorized to download attendance for this subject or session year.")
                return redirect('student_view_attendance')
            attendance_report = Attendance_Report.objects.filter(
                student_id=student,
                attendance_id__subject_id=get_subject,
                attendance_id__session_year_id=get_session_year.id
            ).order_by('attendance_id__attendance_data')
        except (Subject.DoesNotExist, Session_Year.DoesNotExist, ValueError) as e:
            messages.error(request, "Invalid subject or session year selected.")
            return redirect('student_view_attendance')

        if attendance_report.exists():
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Attendance"
            ws.append(["Date", "Status"])
            for report in attendance_report:
                ws.append([report.attendance_id.attendance_data.strftime('%Y-%m-%d'), "Present" if report.status == 1 else "Absent"])
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f'attachment; filename=attendance_{get_subject.name}_{get_session_year.session_start}_to_{get_session_year.session_end}.xlsx'
            wb.save(response)
            print("Excel file generated and sent for student")
            return response
        else:
            messages.error(request, "No attendance data available to export.")
            return redirect('student_view_attendance')

    if 'download_all' in request.GET and request.GET['download_all'] == 'excel':
        session_year_id = request.GET.get('session_year_id')
        print(f"Download All Subjects Excel triggered - session_year_id: {session_year_id}")

        try:
            get_session_year = Session_Year.objects.get(id=int(session_year_id))
            if get_session_year.id != student.session_year_id.id:
                messages.error(request, "You are not authorized to download attendance for this session year.")
                return redirect('student_view_attendance')
            all_subjects = Subject.objects.filter(course=student.course_id)
            attendance_reports = Attendance_Report.objects.filter(
                student_id=student,
                attendance_id__session_year_id=get_session_year.id,
                attendance_id__subject_id__in=all_subjects
            ).order_by('attendance_id__subject_id__name', 'attendance_id__attendance_data')
        except (Session_Year.DoesNotExist, ValueError) as e:
            messages.error(request, "Invalid session year selected.")
            return redirect('student_view_attendance')

        if attendance_reports.exists():
            wb = openpyxl.Workbook()
            wb.remove(wb.active)  # Remove the default sheet

            for subject in all_subjects:
                subject_reports = attendance_reports.filter(attendance_id__subject_id=subject)
                if subject_reports.exists():
                    ws = wb.create_sheet(title=subject.name)
                    ws.append(["Date", "Status"])
                    for report in subject_reports:
                        ws.append([report.attendance_id.attendance_data.strftime('%Y-%m-%d'), "Present" if report.status == 1 else "Absent"])

            if not wb.sheetnames:
                messages.error(request, "No attendance data available to export for any subjects.")
                return redirect('student_view_attendance')

            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f'attachment; filename=attendance_all_subjects_{get_session_year.session_start}_to_{get_session_year.session_end}.xlsx'
            wb.save(response)
            print("Excel file with all subjects generated and sent for student")
            return response
        else:
            messages.error(request, "No attendance data available to export for any subjects.")
            return redirect('student_view_attendance')

    context = {
        'subjects': subjects,
        'session_year': session_year,
        'action': action,
        'get_subject': get_subject,
        'get_session_year': get_session_year,
        'attendance_report': attendance_report,
    }
    return render(request, 'Student/view_attendance.html', context)

@login_required(login_url='/')
def VIEW_RESULT(request):
    student = Student.objects.get(admin=request.user.id)
    results = StudentResult.objects.filter(student_id=student)

    if 'download' in request.GET and request.GET['download'] == 'pdf':
        print("Download PDF triggered for student")
        # Create a PDF using reportlab
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter)
        elements = []

        # Styles
        styles = getSampleStyleSheet()
        title_style = styles['Heading1']
        normal_style = styles['Normal']

        # Add title and student info
        elements.append(Paragraph("Student Results", title_style))
        elements.append(Paragraph(f"Student: {student.admin.first_name} {student.admin.last_name}", normal_style))
        elements.append(Paragraph(f"ID: {student.admin.id}", normal_style))
        elements.append(Paragraph("<br/><br/>", normal_style))  # Spacing

        # Table data
        data = [['Subject', 'Assignment', 'Exam', 'IA1', 'IA2', 'Attendance', 'Mid Sem', 'End Sem', 'CGPA']]
        for result in results:
            data.append([
                result.subject_id.name,
                str(result.assignment_mark) if result.assignment_mark is not None else '-',
                str(result.exam_mark) if result.exam_mark is not None else '-',
                str(result.ia1_mark) if result.ia1_mark is not None else '-',
                str(result.ia2_mark) if result.ia2_mark is not None else '-',
                str(result.attendance_mark) if result.attendance_mark is not None else '-',
                str(result.midsem_mark) if result.midsem_mark is not None else '-',
                str(result.end_sem_mark) if result.end_sem_mark is not None else '-',
                str(result.calculate_cgpa()) if result.calculate_cgpa() is not None else '-'
            ])

        # Create table
        table = Table(data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 10),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        elements.append(table)

        # Build PDF
        doc.build(elements)
        pdf = buffer.getvalue()
        buffer.close()

        # Create response
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename=my_results_{student.admin.username}.pdf'
        response.write(pdf)
        print("PDF file generated and sent for student")
        return response

    context = {'results': results}
    return render(request, 'Student/view_result.html', context)

# New views for student notes
@login_required(login_url='/')
def STUDENT_ASK_DOUBT(request):
    student = Student.objects.get(admin=request.user)
    subjects = Subject.objects.filter(course=student.course_id)
    doubts = StudentDoubt.objects.filter(student=student)

    if request.method == "POST":
        subject_id = request.POST.get('subject_id')
        question = request.POST.get('question')
        
        if not all([subject_id, question]):
            messages.error(request, "Please select a subject and enter your question")
            return redirect('student_ask_doubt')
            
        try:
            subject = Subject.objects.get(id=subject_id)
            
            # Configure Gemini AI
            genai.configure(api_key=settings.GEMINI_API_KEY)
            model = genai.GenerativeModel('gemini-2.0-flash')  # Using the free tier model
            
            prompt = f"""As an educational AI assistant, help with this academic question:
            Subject: {subject.name}
            Course: {subject.course.name}
            Question: {question}
            
            Provide:
            1. Clear explanation
            2. Examples if relevant
            3. Key points to remember
            """
            
            try:
                response = model.generate_content(prompt)
                
                if response.text:
                    ai_response = response.text
                else:
                    raise Exception("Empty response from AI")
                    
            except Exception as e:
                print(f"AI Error: {str(e)}")
                messages.warning(request, "There was an issue getting the AI response. Please try again.")
                return redirect('student_ask_doubt')
            
            # Save the doubt and response
            doubt = StudentDoubt.objects.create(
                student=student,
                subject=subject,
                question=question,
                ai_response=ai_response
            )
            
            messages.success(request, "Your doubt has been answered!")
            return redirect('student_ask_doubt')
            
        except Subject.DoesNotExist:
            messages.error(request, "Invalid subject selected")
        except Exception as e:
            print(f"General Error: {str(e)}")
            messages.error(request, "An error occurred while processing your request. Please try again.")
            
    context = {
        'subjects': subjects,
        'doubts': doubts
    }
    return render(request, 'Student/ask_doubt.html', context)

@login_required(login_url='/')
def STUDENT_VIEW_TEACHER_NOTES(request):
    try:
        student = Student.objects.get(admin=request.user.id)
        # Get subjects for this student's course
        subjects = Subject.objects.filter(course=student.course_id)
        
        # Get selected subject filter
        selected_subject = request.GET.get('subject')
        
        if selected_subject:
            notes = SubjectNote.objects.filter(
                subject_id=selected_subject,
                subject__course=student.course_id
            ).order_by('-created_at')
        else:
            notes = SubjectNote.objects.filter(
                subject__course=student.course_id
            ).order_by('-created_at')
        
        context = {
            'subjects': subjects,
            'notes': notes,
            'selected_subject': selected_subject
        }
        return render(request, 'Student/teacher_notes.html', context)
        
    except Student.DoesNotExist:
        messages.error(request, "Student record not found")
        return redirect('login')
    except Exception as e:
        messages.error(request, f"An error occurred: {str(e)}")
        return redirect('login')

@login_required(login_url='/')
def STUDENT_CREATE_NOTE(request):
    if request.method == "POST":
        title = request.POST.get('title')
        content = request.POST.get('content')
        if title and content:
            note = Note(user=request.user, title=title, content=content)
            note.save()
            messages.success(request, 'Note created successfully!')
            return redirect('student_notes')
        else:
            messages.error(request, 'Title and content are required.')
    return render(request, 'Student/create_note.html')

@login_required(login_url='/')
def STUDENT_EDIT_NOTE(request, note_id):
    try:
        note = Note.objects.get(id=note_id, user=request.user)
    except Note.DoesNotExist:
        messages.error(request, 'Note not found.')
        return redirect('student_notes')
    if request.method == "POST":
        title = request.POST.get('title')
        content = request.POST.get('content')
        if title and content:
            note.title = title
            note.content = content
            note.save()
            messages.success(request, 'Note updated successfully!')
            return redirect('student_notes')
        else:
            messages.error(request, 'Title and content are required.')
    context = {
        'note': note,
    }
    return render(request, 'Student/edit_note.html', context)

@login_required(login_url='/')
def STUDENT_DELETE_NOTE(request, note_id):
    try:
        note = Note.objects.get(id=note_id, user=request.user)
        note.delete()
        messages.success(request, 'Note deleted successfully!')
    except Note.DoesNotExist:
        messages.error(request, 'Note not found.')
    return redirect('student_notes')

@login_required(login_url='/')
def STUDENT_NOTES(request):
    try:
        notes = Note.objects.filter(user=request.user).order_by('-created_at')
        context = {
            'notes': notes
        }
        return render(request, 'Student/notes.html', context)
    except Exception as e:
        messages.error(request, f"An error occurred: {str(e)}")
        return redirect('student_home')